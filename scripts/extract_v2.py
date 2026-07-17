#!/usr/bin/env python3
"""
ERAVisor Extractor V2 — Extracción mejorada con dos modos:
  - Modo rápido: solo regex (fecha, ubicación, víctimas)
  - Modo completo: regex + LLM (deepseek-v4-flash) para clasificación Anexo III

Uso: python3 extract_v2.py [--pais ES] [--modo rapido|completo] [--max N]
"""
import json, csv, os, re, sys, time
from pathlib import Path
from datetime import datetime
import requests, PyPDF2, openpyxl
from openpyxl.styles import Font, PatternFill

BASE = Path("/root/workspace/ERAVisor")
PDF_DIR = BASE / "pdfs"
DATA_DIR = BASE / "data"
NAN_KEY = os.environ.get("NAN_API", "")
LLM_URL = "https://api.nan.builders/v1/chat/completions"
LLM_MODEL = "deepseek-v4-flash"

# Taxonomías
TAX_SUCESOS = {
    "1":"Accidente","1.1":"Colisión tren-vehículo","1.1.1":"Frontal","1.1.2":"Alcance","1.1.3":"Lateral",
    "1.2":"Colisión obstáculo","1.3":"Descarrilamiento","1.3.1":"Plena vía","1.3.2":"Estaciones",
    "1.4":"Accidente paso nivel","1.5":"Atropello persona","1.5.1":"Arrollamiento",
    "1.5.1.1":"Plena vía","1.5.1.2":"Estaciones","1.5.2":"Caídas","1.6":"Incendio/explosión",
    "1.7":"Otros accidentes",
    "2":"Incidente","2.1":"Precursor","2.1.1":"Rotura carril","2.1.3":"Fallo señalización",
    "2.1.4":"Señal rebasada peligro","2.1.6":"Rueda rota","2.1.7":"Eje roto",
    "2.2.1":"Incidente operacional","2.2.1.1":"Exceso velocidad","2.2.1.2":"Marcha no autorizada",
    "2.3":"Otros incidentes",
    "3":"Suicidio","3.1":"Suicidio","3.2":"Intento suicidio",
}
TAX_CAUSAS = {
    "1":"Ferrocarril","1.1":"Factor humano","1.1.1":"Señales","1.1.2":"Bloqueo",
    "1.1.3":"Itinerario","1.1.4":"Formación tren","1.1.5":"Conducción",
    "1.1.6":"Mantenimiento MR","1.1.8":"Maniobras","1.1.10":"Otros",
    "1.2":"Fallo técnico","1.2.1":"Fallo MR","1.2.1.1":"Rodadura","1.2.1.3":"Freno",
    "1.2.1.6":"Pantógrafo","1.2.1.12":"Otros MR",
    "1.2.2":"Fallo instalaciones","1.2.2.2":"Vía","1.2.2.3":"Carril","1.2.2.4":"Aparato vía",
    "1.2.2.5":"Seguridad","1.2.2.6":"Electrificación","1.2.2.7":"Otros",
    "2":"Usuarios/entorno","2.1":"Usuarios FC","2.2":"Condiciones entorno",
    "2.2.1":"Meteorológicas","2.3":"Otros","2.3.3":"Usuario PN","2.4":"Sin identificar",
}

def extraer_regex(texto, anio, pais, pdf_url, titulo):
    """Extracción mejorada con regex — maneja formatos antiguos y modernos"""
    r = {"pais": pais, "titulo_informe": titulo, "informe_publico_url": pdf_url,
         "fecha_procesamiento": datetime.now().isoformat(), "fuente_datos": "ERA",
         "organismo_investigador": "CIAF", "estado_investigacion": "Completado",
         "num_paginas": 0, "fallecidos": 0, "heridos_graves": 0, "heridos_leves": 0}
    
    txt_head = texto[:5000]
    txt_upper = txt_head.upper()
    
    # === FECHA ===
    # Buscar fecha del accidente (no la del informe)
    m = re.search(r'(?:ocurrido|producido|acontecido|acaecido)\s+(?:el día|el)\s+(\d{1,2})[./](\d{1,2})[./.]?(\d{4})', txt_head, re.I)
    if not m:
        m = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', txt_head)
    if m:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        if 1900 < int(y) < 2030 and 1 <= int(mo) <= 12 and 1 <= int(d) <= 31:
            r["fecha"] = f"{y}-{mo}-{d}"
    if "fecha" not in r:
        r["fecha"] = f"{anio}-01-01"
    
    # === HORA ===
    m = re.search(r'(?:a las|sobre las|hacia las|siendo las)\s*(\d{1,2})[:h.](\d{2})?', txt_head, re.I)
    if m:
        h = m.group(1).zfill(2)
        mn = m.group(2) or "00"
        if int(h) < 24:
            r["hora"] = f"{h}:{mn}"
    
    # === Nº INFORME ===
    # De ID en nombre de archivo
    m_id = re.search(r'[nN]°?\s*(\d+)[/\s]*(\d{4})?', txt_head)
    if m_id:
        r["id_accidente"] = f"{pais}-{m_id.group(2) or anio}-{m_id.group(1)}"
    else:
        stem = Path(titulo).stem.replace("ID-","").replace("IF-","").replace("RS-","")
        r["id_accidente"] = f"{pais}-{anio}-{stem[:20]}"
    
    # === PROVINCIA ===
    # Buscar en texto: provincia/cl localidad
    prov_patterns = [
        r'(?:provincia de|en la provincia|localidad de)\s+([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúññ]+(?:\.|\s))?\w+)',
        r'(?:término municipal|municipio|Ayuntamiento)\s+(?:de\s+)?([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)*)',
    ]
    for pat in prov_patterns:
        m = re.search(pat, txt_head, re.I)
        if m:
            prov = m.group(1).strip().rstrip('.')
            # Filtrar términos no geográficos
            if not any(x in prov.lower() for x in ['adif','renfe','ferrocarril','ffe','cercanías']):
                r["provincia"] = prov
                break
    
    # === MUNICIPIO ===
    m = re.search(r'(?:entre|en)\s+(?:las\s+)?estac(?:ión|iones)\s+(?:de\s+)?([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ]+)', txt_head)
    if m:
        r["municipio"] = m.group(1).strip()
    # También buscar estación concreta
    m = re.search(r'(?:estación de|estación)\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑa-záéíóúñ]+)*)', txt_head, re.I)
    if m and "municipio" not in r:
        r["municipio"] = m.group(1).strip().split(',')[0]
    
    # === LÍNEA ===
    m = re.search(r'(?:línea|línea|Línea)\s+(?:\d+\s+)?([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑa-záéíóúñ]+){0,4}(?:[-–]\s*[A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ]+)?)', txt_head)
    if m:
        ln = m.group(1).strip()
        if len(ln) > 3 and ln not in ['de','del','la','las','los','y']:
            r["linea"] = ln
    else:
        m = re.search(r'LÍNEA\s+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ]+)', txt_head, re.I)
        if m:
            r["linea"] = m.group(1).strip()
    
    # === PK ===
    m = re.search(r'(?:PK|p\.k\.|punto kilométrico|km)\s*[.:]?\s*(\d+)[+,.](\d+)', txt_head, re.I)
    if m:
        r["pk"] = f"{m.group(1)}+{m.group(2).zfill(3)}"
    else:
        m = re.search(r'(?:PK|km)\s*[.:]?\s*(\d{2,4})', txt_head, re.I)
        if m:
            r["pk"] = m.group(1)
    
    # === OPERADOR ===
    operadores = {"RENFE":"Renfe","ADIF":"ADIF","FEVE":"FEVE","Renfe Viajeros":"Renfe","Renfe Mercancías":"Renfe"}
    for op, norm in operadores.items():
        if op in txt_head[:3000]:
            r["operador"] = norm
            break
    if "operador" not in r:
        m = re.search(r'(?:empresa|operador)\s*(?:ferroviaria|ferroviario)?\s*(?::|de)?\s*([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ]+)', txt_head, re.I)
        if m:
            r["operador"] = m.group(1)
    
    # === VÍCTIMAS ===
    # Buscar números específicos con contextos variados
    victim_patterns = [
        ("fallecidos", r'(\d+)\s*(?:fallecido[s]?|víctima[s]? mortal[es]?|muerto[s]?|fallecida[s]?)'),
        ("fallecidos", r'(?:falleci[oó]|muri[oó]|result[oó] fallecido)\s*[.,]?\s*(\d+)'),
        ("heridos_graves", r'(\d+)\s*(?:herido[s]? grave[s]?|herida[s]? grave[s]?)'),
        ("heridos_leves", r'(\d+)\s*(?:herido[s]? leve[s]?|herida[s]? leve[s]?)'),
        ("heridos_graves", r'(\d+)\s*(?:grave[s]?)\s*(?:herido|lesionado)'),
    ]
    for key, pat in victim_patterns:
        if key in r and r.get(key,0) > 0:
            continue
        m = re.search(pat, txt_head, re.I)
        if m:
            val = int(m.group(1))
            if val < 500:  # sanity check
                r[key] = val
    
    # === TIPO DE SUCESO (clasificación jerárquica mejorada) ===
    # Orden: más específico primero
    suceso = ("1.7", "Otros accidentes")
    
    if re.search(r'DESCARRILAMIENTO|DESCARRIL[OÓ]|SALIDA DE VIA|SALIDA DE VÍA', txt_upper):
        if re.search(r'ESTACION|ESTACIÓN|ANDÉN|ANDE[N]', txt_upper):
            suceso = ("1.3.2", "Descarrilamiento en estación")
        else:
            suceso = ("1.3.1", "Descarrilamiento en plena vía")
    elif re.search(r'COLISI[OÓ]N.*ALCANCE|ALCANCE.*COLISI', txt_upper):
        suceso = ("1.1.2", "Colisión por alcance")
    elif re.search(r'COLISI[OÓ]N.*FRONTAL|FRONTAL.*COLISI', txt_upper):
        suceso = ("1.1.1", "Colisión frontal")
    elif re.search(r'COLISI[OÓ]N', txt_upper):
        suceso = ("1.1", "Colisión")
    elif re.search(r'PASO\s+(?:A\s+)?NIVEL', txt_upper):
        if re.search(r'VEH[IÍ]CULO|COCHE|AUTOM[OÓ]VIL|CAMION', txt_upper):
            suceso = ("1.4.1.1", "Accidente PN - Vehículo")
        elif re.search(r'PEAT[OÓ]N|PEATONES', txt_upper):
            suceso = ("1.4.1.2", "Accidente PN - Peatón")
        else:
            suceso = ("1.4", "Accidente en paso a nivel")
    elif re.search(r'ARROLLAMIENTO|ARROLLO|ATROPELLO|ATROPELLAMIENTO|PERSONA.*ARR[OÓ]L', txt_upper):
        if re.search(r'ESTACION|ESTACIÓN|ANDE[N]', txt_upper):
            suceso = ("1.5.1.2", "Arrollamiento en estación")
        else:
            suceso = ("1.5.1.1", "Arrollamiento en plena vía")
    elif re.search(r'INCENDIO|EXPLOSI[OÓ]N|FUEGO', txt_upper):
        suceso = ("1.6", "Incendio/explosión")
    elif re.search(r'CA[IÍ]DA|CAERSE|CA[IÍ]DO|RESBAL[OÓ]N|CAIDA', txt_upper):
        suceso = ("1.5.2", "Caída de persona")
    elif re.search(r'SUICIDIO|SUICIDA|INTENTO DE SUICIDIO', txt_upper):
        if re.search(r'INTENTO', txt_upper):
            suceso = ("3.2", "Intento de suicidio")
        else:
            suceso = ("3.1", "Suicidio")
    elif re.search(r'SE[ÑN]AL.*REBASA|REBASA.*SE[ÑN]AL|REBASAMIENTO|SPAD', txt_upper):
        suceso = ("2.1.4", "Señal rebasada")
    elif re.search(r'INCIDENTE|CASI COLISI[OÓ]N|CONATO', txt_upper):
        if re.search(r'ROTURA\s+DE\s+CARRIL|CARRIL\s+ROTO|ROTURA\s+CARRIL', txt_upper):
            suceso = ("2.1.1", "Rotura de carril")
        elif re.search(r'EXCESO.*VELOCIDAD|VELOCIDAD.*EXCESO', txt_upper):
            suceso = ("2.2.1.1", "Exceso de velocidad")
        elif re.search(r'RUEDA\s+ROTA|ROTURA\s+RUEDA|EJE\s+ROTO', txt_upper):
            suceso = ("2.1.6/7", "Rueda/eje roto en servicio")
        else:
            suceso = ("2", "Incidente")
    else:
        suceso = ("1.7.7", "Otros")
    
    r["suceso_codigo"], r["suceso_desc"] = suceso
    
    # === CAUSA ===
    causa = ("2.4", "Sin identificar")
    
    if re.search(r'FACTOR\s+HUMANO|ERROR\s+HUMANO|CONDUCTOR.*ERROR|MAQUINISTA.*ERROR|CONDUCCI[OÓ]N.*INADECUADA', txt_upper):
        if re.search(r'SE[ÑN]AL|SEÑALES|BLOQUEO|ITINERARIO', txt_upper):
            causa = ("1.1.1", "Factor humano - Señales")
        elif re.search(r'EXCESO.*VELOCIDAD|VELOCIDAD.*EXCESO', txt_upper):
            causa = ("1.1.5", "Factor humano - Exceso velocidad")
        else:
            causa = ("1.1.5", "Factor humano - Conducción")
    elif re.search(r'FALLO.*FRENO|FRENO.*FALLO|ROTURA.*FRENO|FRENO.*ROTO', txt_upper):
        causa = ("1.2.1.3", "Fallo técnico - Freno")
    elif re.search(r'ROTURA.*CARRIL|CARRIL.*ROTO|DEFORMACI[OÓ]N.*CARRIL|PANDEO', txt_upper):
        causa = ("1.2.2.3", "Fallo instalaciones - Carril")
    elif re.search(r'FALLO.*VIA|VÍA.*FALLO|DESALINEACI[OÓ]N.*VIA|VIA.*DEFORM', txt_upper):
        causa = ("1.2.2.2", "Fallo instalaciones - Vía")
    elif re.search(r'FALLO.*SE[ÑN]ALIZACI[OÓ]N|SE[ÑN]AL.*FALLA|SEÑAL.*FALLA', txt_upper):
        causa = ("1.2.2.5", "Fallo instalaciones - Seguridad")
    elif re.search(r'ROTURA.*RUEDA|RUEDA.*ROTA|ROTURA.*EJE|EJE.*ROTO|DETECTOGA|RUEDA.*DEFORM', txt_upper):
        causa = ("1.2.1.1", "Fallo técnico - Rodadura")
    elif re.search(r'METEOROL[OÓ]GICA|LLUVIA.*INTENSA|VIENTO.*FUERTE|NIEBLA|TORMENTA|HUE[Ll]O', txt_upper):
        causa = ("2.2.1", "Condiciones meteorológicas")
    elif re.search(r'USUARIO.*PASO.*NIVEL|PEAT[OÓ]N.*PASO.*NIVEL|VEH[IÍ]CULO.*PASO.*NIVEL', txt_upper):
        causa = ("2.3.3", "Usuario paso a nivel")
    elif re.search(r'INTRUSO|EXTRA[ÑN]O|VANDALISMO', txt_upper):
        causa = ("2.3.2", "Intrusión")
    elif re.search(r'VIAJERO.*CA[IÍ]DA|VIAJERO.*SUBIR|VIAJERO.*BAJAR|PASAJERO.*CA[IÍ]DA', txt_upper):
        causa = ("2.1.1", "Usuario ferroviario - Viajero")
    elif re.search(r'MANTENIMIENTO', txt_upper):
        causa = ("1.1.6", "Factor humano - Mantenimiento")
    
    r["causa_codigo"], r["causa_desc"] = causa
    
    # === RESUMEN (primer párrafo útil) ===
    lines = [l.strip() for l in texto.split('\n') if l.strip() and len(l.strip()) > 30 and not l.strip().startswith('SECRETARIA') and 'INFRAESTRUCTURAS' not in l.upper()]
    if lines:
        r["resumen"] = lines[0][:500]
    
    return r


def extraer_llm(texto, anio, pais):
    """Extracción con LLM para clasificación precisa"""
    txt = texto[:10000]
    prompt = f"""Eres un experto en seguridad ferroviaria. Analiza este informe de accidente ferroviario y extrae los datos estructurados.

INFORME:
{txt}

Responde SOLO con este JSON exacto (sin texto adicional):
{{"fecha":"YYYY-MM-DD","hora":"HH:MM","provincia":"Madrid","municipio":"","operador":"Renfe","infraestructura":"ADIF","linea":"Madrid-Hendaya","pk":"73+885","tipo_trafico":"mercancias|viajeros|mixto|maniobras","tipo_linea":"convencional|alta_velocidad|ancho_metrico","paso_nivel":false,"estacion":false,"suceso_codigo":"1.1.2","suceso_desc":"Colision por alcance","causa_codigo":"1.1.5","causa_desc":"Factor humano - Conduccion","fallecidos":0,"heridos_graves":0,"heridos_leves":0,"trenes":2,"resumen":"Breve resumen del accidente","causas_directas":"Causas directas identificadas","factores_contribuyentes":"Factores contribuyentes","recomendaciones":"Recomendaciones de seguridad"}}

Usa null si no puedes determinar un campo. Codigos de suceso: 1.1(colision), 1.1.1(frontal), 1.1.2(alcance), 1.1.3(lateral), 1.2(colision obstaculo), 1.3(descarrilamiento), 1.3.1(plena via), 1.3.2(estaciones), 1.4(paso nivel), 1.5(atropello), 1.5.1(arrollamiento), 1.5.1.1(plena via), 1.5.1.2(estaciones), 1.5.2(caidas), 1.6(incendio), 2(incidente), 2.1.1(rotura carril), 2.1.3(fallo senalizacion), 2.1.4(senal rebasada), 2.1.6(rueda rota), 2.1.7(eje roto), 2.2.1.1(exceso velocidad), 2.2.1.2(marcha no autorizada), 3(suicidio)
Codigos causa: 1.1(factor humano), 1.1.1(senales), 1.1.2(bloqueo), 1.1.5(conduccion), 1.2(fallo tecnico), 1.2.1.1(rodadura), 1.2.1.3(freno), 1.2.1.6(pantografo), 1.2.2.2(via), 1.2.2.3(carril), 1.2.2.4(aparato via), 1.2.2.5(seguridad), 2.2.1(meteorologico), 2.3.3(usuario PN), 2.4(sin identificar)"""
    
    try:
        r = requests.post(LLM_URL, headers={
            "Authorization": f"Bearer {NAN_KEY}", "Content-Type": "application/json"
        }, json={
            "model": LLM_MODEL, "messages": [
                {"role": "system", "content": "Eres un experto en seguridad ferroviaria. Responde SOLO con JSON válido."},
                {"role": "user", "content": prompt}
            ], "temperature": 0.05, "max_tokens": 1500
        }, timeout=120)
        if r.status_code == 200:
            text = r.json()["choices"][0]["message"]["content"]
            m = re.search(r'\{.*\}', text, re.DOTALL)
            if m:
                return json.loads(m.group())
    except:
        pass
    return None


def procesar(report, modo, pais_dir):
    title = report["title"]
    pdf_url = report["pdf_url"]
    year = report["year"]
    country = report["country"]
    fname = f"{country}_{year}_{title}"
    path = pais_dir / fname
    
    if not path.exists():
        print(f"  ⚠️  No descargado: {title}")
        return None
    
    try:
        with open(path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            paginas = []
            for p in reader.pages:
                paginas.append(p.extract_text())
            texto = "\n\n".join(paginas)
            npag = len(reader.pages)
    except:
        print(f"  ❌ Error leyendo PDF: {title}")
        return None
    
    # Extracción base con regex
    datos = extraer_regex(texto, year, country, pdf_url, title)
    datos["num_paginas"] = npag
    
    if modo == "completo":
        print(f"  🤖 LLM...", end=" ", flush=True)
        llm_data = extraer_llm(texto, year, country)
        if llm_data:
            for k, v in llm_data.items():
                if v is not None:
                    datos[k] = v
            print("✅")
        else:
            print("⚠️ fallback regex")
    else:
        print(f"  📋 regex OK")
    
    return datos


def generar_excel_csv(resultados, pais):
    if not resultados:
        return
    keys = sorted(set().union(*(r.keys() for r in resultados)))
    
    # CSV
    csv_path = DATA_DIR / f"{pais.lower()}_accidentes.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        w.writerows(resultados)
    print(f"\n📄 CSV: {csv_path} ({len(resultados)} rows, {len(keys)} cols)")
    
    # Excel
    xlsx_path = DATA_DIR / f"{pais.lower()}_accidentes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Accidentes {pais}"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for c, k in enumerate(keys, 1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.font = hf
        cell.fill = hfill
    for ri, res in enumerate(resultados, 2):
        for c, k in enumerate(keys, 1):
            ws.cell(row=ri, column=c, value=res.get(k, ""))
    wb.save(xlsx_path)
    print(f"📊 Excel: {xlsx_path} ({len(resultados)} rows, {len(keys)} cols)")


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--pais", default="ES")
    parser.add_argument("--modo", choices=["rapido", "completo"], default="rapido")
    parser.add_argument("--max", type=int, default=0)
    args = parser.parse_args()
    
    idx_file = DATA_DIR / f"{args.pais.lower()}-investigations-index.json"
    with open(idx_file) as f:
        reports = json.load(f)["reports"]
    
    pais_dir = PDF_DIR / args.pais
    print(f"🚆 ERAVisor V2 — {args.pais} | {args.modo} | {len(reports)} reports")
    
    if args.max:
        reports = reports[:args.max]
    
    resultados = []
    for i, r in enumerate(reports):
        print(f"[{i+1}/{len(reports)}] {r['title'][:60]}", end=" ", flush=True)
        d = procesar(r, args.modo, pais_dir)
        if d:
            resultados.append(d)
    
    if resultados:
        generar_excel_csv(resultados, args.pais)
        json_path = DATA_DIR / f"{args.pais.lower()}_extraido.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(resultados, f, ensure_ascii=False, indent=2)
        print(f"💾 JSON: {json_path}")
    
    print(f"\n✅ {len(resultados)}/{len(reports)} procesados")

if __name__ == "__main__":
    main()