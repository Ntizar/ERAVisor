#!/usr/bin/env python3
"""
ERAVisor — Pipeline Europa (phase 2)
  Usa los índices existentes para descargar PDFs y extraer datos.
  No hace scraping de ERA (Drupal 11 JS).
  Se ejecuta cada 10 minutos vía cron.

Fases por ejecución (rotativas):
  Fase 1: Descargar PDFs (países sin PDFs)
  Fase 2: Extraer con LLM (países con PDFs pero sin extraer)
  Fase 3: Actualizar visor
"""
import json, csv, os, re, sys, time
from pathlib import Path
from datetime import datetime
import requests

BASE = Path("/root/workspace/ERAVisor")
PDF_DIR = BASE / "pdfs"
DATA_DIR = BASE / "data"
VISOR_DIR = BASE / "visor"

NAN_KEY = os.environ.get("NAN_API", "")
LLM_URL = "https://api.nan.builders/v1/chat/completions"
LLM_MODEL = "qwen3.6"

# === TAXONOMÍA COMPLETA Anexo III ===
TAX_SUCESOS = {
    "1":"Accidente","1.1":"Colisión tren-vehículo","1.1.1":"Frontal","1.1.2":"Alcance","1.1.3":"Lateral",
    "1.2":"Colisión obstáculo","1.2.1":"Elementos tren","1.2.2":"Superestructura","1.2.3":"Infraestructura",
    "1.2.4":"Protección itinerarios","1.2.5":"Animales","1.2.6":"Árboles","1.2.7":"Maq. mantenimiento","1.2.8":"Otros",
    "1.3":"Descarrilamiento","1.3.1":"Plena vía","1.3.2":"Estaciones",
    "1.4":"Accidente paso nivel","1.4.1":"PN Usuarios","1.4.1.1":"PN Vehículo","1.4.1.2":"PN Peatón","1.4.2":"PN Objetos",
    "1.5":"Atropello persona","1.5.1":"Arrollamiento","1.5.1.1":"Plena vía","1.5.1.2":"Estaciones",
    "1.5.1.2.1":"Cruce andenes","1.5.1.2.2":"Estaciones otros","1.5.2":"Caídas","1.5.2.1":"En marcha","1.5.2.2":"Subir/bajar",
    "1.5.3":"Otros personas","1.5.3.1":"Movimiento vehículo","1.5.3.2":"Proyección objetos","1.5.3.3":"Otros",
    "1.6":"Incendio/explosión",
    "1.7":"Otros accidentes","1.7.1":"Proc. emergencia","1.7.2":"Maniobras","1.7.3":"Cambio ancho",
    "1.7.4":"Vía bloqueada","1.7.5":"Mercancías peligrosas","1.7.6":"Fallo cargamento","1.7.7":"Otros",
    "2":"Incidente","2.1":"Precursor","2.1.1":"Rotura carril","2.1.2":"Deformación vía",
    "2.1.3":"Fallo señalización","2.1.3.1":"Señalización infra","2.1.3.2":"Señalización MR",
    "2.1.4":"SPAD sobrepasando","2.1.4.1":"Conato colisión","2.1.4.2":"SPAD otros",
    "2.1.5":"SPAD sin sobrepasar","2.1.6":"Rueda rota","2.1.7":"Eje roto",
    "2.2":"Otros precursores","2.2.1":"Incidentes operacionales",
    "2.2.1.1":"Exceso velocidad","2.2.1.2":"Marcha no autorizada","2.2.1.3":"Incumplimiento bloqueo","2.2.1.4":"Otros incumplimientos",
    "2.2.2":"Escape material","2.2.2.1":"Conato colisión","2.2.2.2":"Escape otros",
    "2.3":"Otros incidentes","2.3.1":"Componentes vehículo","2.3.2":"Fallo cargamento","2.3.3":"Incidente PN",
    "2.3.4":"Talonamiento","2.3.5":"Conato incendio","2.3.6":"Mercancías peligrosas",
    "2.3.7":"Incidente obstáculo gálibo","2.3.7.1":"Elementos tren","2.3.7.2":"Superestructura","2.3.7.3":"Infraestructura",
    "2.3.7.4":"Protección itinerarios","2.3.7.5":"Animales","2.3.7.6":"Árboles","2.3.7.7":"Otros","2.3.8":"Otras incidencias",
    "3":"Suicidio","3.1":"Suicidio","3.2":"Intento suicidio",
}
TAX_CAUSAS = {
    "1":"Ferrocarril","1.1":"Factor humano","1.1.1":"Señales","1.1.2":"Bloqueo","1.1.3":"Itinerario",
    "1.1.4":"Formación tren","1.1.5":"Conducción","1.1.6":"Mantenimiento","1.1.7":"Trabajos vía",
    "1.1.8":"Maniobras","1.1.9":"Prescripciones","1.1.10":"Otros FH",
    "1.2":"Fallo técnico","1.2.1":"Fallo MR","1.2.1.1":"Rodadura","1.2.1.2":"Suspensión","1.2.1.3":"Freno",
    "1.2.1.4":"Tracción-choque","1.2.1.5":"Gálibo","1.2.1.6":"Pantógrafo","1.2.1.7":"Puertas",
    "1.2.1.8":"Seguridad embarcada","1.2.1.9":"Motor eléctrico","1.2.1.10":"Motor combustión",
    "1.2.1.11":"Elemento caído","1.2.1.12":"Otros MR",
    "1.2.2":"Fallo instalaciones","1.2.2.1":"Infraestructura","1.2.2.2":"Vía","1.2.2.3":"Carril",
    "1.2.2.4":"Aparato vía","1.2.2.5":"Seguridad","1.2.2.6":"Electrificación","1.2.2.7":"Otras instalaciones",
    "2":"Usuarios/entorno","2.1":"Usuarios FC","2.1.1":"Viajeros","2.1.2":"Cliente mercancías","2.1.3":"Otros usuarios",
    "2.2":"Condiciones entorno","2.2.1":"Meteorológicas","2.2.2":"Medio ambiente",
    "2.3":"Otros","2.3.1":"Extraños","2.3.2":"Intrusos","2.3.3":"Usuario PN","2.3.4":"Personas andén","2.3.5":"Personas fuera andén",
    "2.4":"Sin identificar",
}

# === Estados de progreso (persistidos en disco) ===
STATE_FILE = DATA_DIR / "pipeline_state.json"

def cargar_estado():
    if STATE_FILE.exists():
        return json.load(open(STATE_FILE))
    return {
        "downloaded": {},    # {"DE": {"title": true, ...}}
        "extracted": {},     # {"DE": {"title": true, ...}}
        "last_update": None
    }

def guardar_estado(state):
    state["last_update"] = datetime.now().isoformat()
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)


# === Cargar índices ===
def cargar_indices():
    """Carga todos los índices existentes"""
    indices = {}
    for idx_file in DATA_DIR.glob("*-investigations-index.json"):
        pais = idx_file.stem.split("-")[0].upper()
        try:
            with open(idx_file) as f:
                data = json.load(f)
            reports = data.get("reports", [])
            indices[pais] = reports
        except:
            pass
    return indices


# === DESCARGA DE PDFs ===
def descargar_pdf(pdf_url, dest_path):
    """Descarga un PDF con reintentos"""
    for intento in range(3):
        try:
            resp = requests.get(pdf_url, headers={
                "User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0"
            }, timeout=60, allow_redirects=True)
            if resp.status_code == 200 and resp.content[:5] == b"%PDF-":
                dest_path.write_bytes(resp.content)
                return True
            elif resp.status_code == 429:
                time.sleep(3 * (intento+1))
            else:
                break
        except:
            time.sleep(2 * (intento+1))
    return False


def descargar_pendientes(indices, state, batch_size=10):
    """Descarga PDFs pendientes de un país (batch para no saturar)"""
    total_descargados = 0
    total_pendientes = 0
    
    for pais, reports in indices.items():
        pais_dir = PDF_DIR / pais
        pais_dir.mkdir(parents=True, exist_ok=True)
        
        downloaded = state.get("downloaded", {}).get(pais, {})
        pendientes = []
        
        for r in reports:
            title = r["title"]
            if title in downloaded and downloaded[title]:
                continue
            
            fname = f"{pais}_{r['year']}_{title}"
            path = pais_dir / fname
            
            if path.exists() and path.stat().st_size > 2000:
                downloaded[title] = True
                total_descargados += 1
                continue
            
            pendientes.append((r, path))
        
        if not pendientes:
            continue
        
        total_pendientes += len(pendientes)
        
        # Descargar en batch
        for i, (r, path) in enumerate(pendientes[:batch_size]):
            title = r["title"]
            url = r["pdf_url"]
            print(f"  📥 [{pais}] {title[:55]}", end=" ", flush=True)
            
            if descargar_pdf(url, path):
                downloaded[title] = True
                total_descargados += 1
                print("✅")
            else:
                print("❌")
            
            time.sleep(1.5)  # Delay entre descargas
    
    return total_descargados, total_pendientes


# === EXTRACCIÓN CON LLM ===
def clasificar_con_qwen(texto, pais):
    """Clasifica el accidente usando qwen3.6"""
    prompt = f"""Eres un experto en seguridad ferroviaria. Analiza este informe de accidente y clasifícalo según el Anexo III del RD 929/2020.

INFORME (país: {pais}):
{texto[:6000]}

CLASIFICACIÓN DISPONIBLE (código - descripción):
SUCESOS: {', '.join(f'{k}={v}' for k,v in sorted(TAX_SUCESOS.items()))}
CAUSAS: {', '.join(f'{k}={v}' for k,v in sorted(TAX_CAUSAS.items()))}

Responde SOLO con este JSON exacto (sin texto adicional, sin markdown):
{{"suceso_codigo":"1.1.1","suceso_desc":"Colisión frontal","causa_codigo":"1.1.5","causa_desc":"FH - Conducción","provincia":"","municipio":"","operador":"","linea":"","fecha":"","hora":"","pk":"","fallecidos":0,"heridos_graves":0,"heridos_leves":0,"resumen":"Descripción breve del accidente"}}

Usa código de suceso y causa de la lista. Si no puedes determinar, usa "1.7.7" para suceso y "2.4" para causa. Responde SÓLO el JSON."""
    
    try:
        r = requests.post(LLM_URL, headers={
            "Authorization": f"Bearer {NAN_KEY}", "Content-Type": "application/json"
        }, json={
            "model": LLM_MODEL,
            "messages": [
                {"role": "system", "content": "Eres un experto en seguridad ferroviaria. Responde SOLO con JSON válido."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.05,
            "max_tokens": 1200
        }, timeout=180)
        
        if r.status_code == 200:
            text = r.json()["choices"][0]["message"]["content"]
            m = re.search(r'\{.*\}', text, re.DOTALL)
            if m:
                return json.loads(m.group())
    except Exception as e:
        print(f"⚠️ LLM error: {e}", end=" ")
    
    return None


def extraer_pdf(path_pdf, title, pais, year, pdf_url):
    """Extrae datos de un PDF"""
    import PyPDF2
    try:
        reader = PyPDF2.PdfReader(open(path_pdf, "rb"))
        paginas = [p.extract_text() or "" for p in reader.pages]
        texto = "\n\n".join(paginas)
        npag = len(reader.pages)
    except Exception as e:
        print(f"  ❌ Error PDF: {e}")
        return None
    
    txt_head = texto[:5000]
    txt_upper = txt_head.upper()
    
    # Organismos por país
    organismos = {
        "ES": "CIAF", "DE": "BAF", "FR": "BEF", "IT": "ANSF",
        "UK": "RAIB", "NL": "OV", "AT": "BSB", "BE": "BEF",
        "CH": "BFTS", "PL": "UIR", "PT": "INR", "NO": "TSB"
    }
    
    datos = {
        "pais": pais,
        "titulo_informe": title,
        "informe_publico_url": pdf_url,
        "fecha_procesamiento": datetime.now().isoformat(),
        "fuente_datos": "ERA",
        "organismo_investigador": organismos.get(pais, pais),
        "estado_investigacion": "Completado",
        "num_paginas": npag,
        "fallecidos": 0, "heridos_graves": 0, "heridos_leves": 0,
        "latitud": 48.0, "longitud": 10.0,  # Centro Europa por defecto
    }
    
    # Extraer fecha
    m = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', txt_head)
    if m and 1900 < int(m.group(3)) < 2030:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        datos["fecha"] = f"{y}-{mo}-{d}"
    else:
        datos["fecha"] = f"{year}-01-01"
    
    # Generar ID
    stem = Path(title).stem[:20]
    datos["id_accidente"] = f"{pais}-{year}-{stem}"
    
    # Clasificación con qwen3.6
    print(f"🤖", end=" ", flush=True)
    llm_data = clasificar_con_qwen(texto[:8000], pais)
    if llm_data:
        for k in ["suceso_codigo", "suceso_desc", "causa_codigo", "causa_desc", "resumen",
                   "provincia", "municipio", "operador", "linea", "pk", "hora",
                   "fallecidos", "heridos_graves", "heridos_leves"]:
            if k in llm_data and llm_data[k] not in (None, "", "null"):
                datos[k] = llm_data[k]
        print("✅")
    else:
        print("⚠️ fallback")
        # Fallback regex multilingüe
        if any(x in txt_upper for x in ["DESCARRILAMIENTO", "DERAILMENT", "DERAILLEMENT", "ENTGLEISEN"]):
            datos["suceso_codigo"] = "1.3"
            datos["suceso_desc"] = TAX_SUCESOS["1.3"]
        elif any(x in txt_upper for x in ["COLISI", "COLLISION", "ZUSAMMENSTOß", "BOTTE"]):
            datos["suceso_codigo"] = "1.1"
            datos["suceso_desc"] = TAX_SUCESOS["1.1"]
        elif any(x in txt_upper for x in ["PASO A NIVEL", "LEVEL CROSSING", "PASSE À NIVEAU", "BAHNÜBERGANG"]):
            datos["suceso_codigo"] = "1.4"
            datos["suceso_desc"] = TAX_SUCESOS["1.4"]
        elif any(x in txt_upper for x in ["ARROLLAMIENTO", "ATROPELLO", "RUN OVER", "ÜBERFAHREN"]):
            datos["suceso_codigo"] = "1.5"
            datos["suceso_desc"] = TAX_SUCESOS["1.5"]
        elif any(x in txt_upper for x in ["INCENDIO", "FIRE", "FEU", "BRAND"]):
            datos["suceso_codigo"] = "1.6"
            datos["suceso_desc"] = TAX_SUCESOS["1.6"]
        elif any(x in txt_upper for x in ["SUICIDIO", "SUICIDE", "SUICIDE"]):
            datos["suceso_codigo"] = "3"
            datos["suceso_desc"] = TAX_SUCESOS["3"]
        else:
            datos["suceso_codigo"] = "1.7.7"
            datos["suceso_desc"] = TAX_SUCESOS["1.7.7"]
        datos["causa_codigo"] = "2.4"
        datos["causa_desc"] = TAX_CAUSAS["2.4"]
    
    # Resumen
    lines = [l.strip() for l in texto.split('\n') if l.strip() and len(l) > 50]
    if lines:
        datos["resumen"] = datos.get("resumen", lines[0][:400])
    
    return datos


def extraer_pendientes(indices, state, batch_size=5):
    """Extrae con LLM PDFs que ya tienen descargados pero no extraídos"""
    total_extraidos = 0
    total_pendientes = 0
    
    for pais, reports in indices.items():
        pais_dir = PDF_DIR / pais
        extraido_file = DATA_DIR / f"{pais.lower()}_extraido.json"
        
        # Cargar extraídos previos
        existentes = {}
        if extraido_file.exists():
            for d in json.load(open(extraido_file)):
                if "titulo_informe" in d:
                    existentes[d["titulo_informe"]] = d
        
        # Descargados
        downloaded = state.get("downloaded", {}).get(pais, {})
        extracted = state.get("extracted", {}).get(pais, {})
        
        pendientes = []
        for r in reports:
            title = r["title"]
            if title in extracted and extracted[title]:
                continue
            
            fname = f"{pais}_{r['year']}_{title}"
            path = pais_dir / fname
            
            if path.exists() and path.stat().st_size > 2000:
                pendientes.append((r, path))
        
        if not pendientes:
            continue
        
        total_pendientes += len(pendientes)
        
        # Extraer en batch
        for i, (r, path) in enumerate(pendientes[:batch_size]):
            title = r["title"]
            print(f"  [{i+1}/{len(pendientes[:batch_size])}] {title[:55]}", end=" ", flush=True)
            
            d = extraer_pdf(path, title, pais, r["year"], r["pdf_url"])
            if d:
                existentes[title] = d
                extracted[title] = True
                total_extraidos += 1
            else:
                print("❌")
                continue
            print("✅", end=" ", flush=True)
            
            time.sleep(1)  # Delay entre LLM calls
        
        # Guardar resultados intermedios
        if existentes:
            resultados = list(existentes.values())
            exportar(resultados, pais)
    
    return total_extraidos, total_pendientes


def exportar(resultados, pais):
    """Exporta a CSV, Excel y JSON"""
    if not resultados:
        return
    
    campos_base = [
        "pais", "id_accidente", "fecha", "municipio", "provincia", "linea", "operador",
        "suceso_codigo", "suceso_desc", "causa_codigo", "causa_desc",
        "fallecidos", "heridos_graves", "heridos_leves",
        "hora", "pk", "resumen", "titulo_informe", "informe_publico_url",
        "organismo_investigador", "num_paginas", "fuente_datos",
        "latitud", "longitud", "estado_investigacion", "fecha_procesamiento"
    ]
    
    # CSV
    csv_path = DATA_DIR / f"{pais.lower()}_accidentes.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=campos_base, extrasaction="ignore")
        w.writeheader()
        w.writerows(resultados)
    
    # Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        xlsx_path = DATA_DIR / f"{pais.lower()}_accidentes.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Accidentes {pais}"
        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        for c, k in enumerate(campos_base, 1):
            cell = ws.cell(row=1, column=c, value=k)
            cell.font = hf
            cell.fill = hfill
        for ri, res in enumerate(resultados, 2):
            for c, k in enumerate(campos_base, 1):
                ws.cell(row=ri, column=c, value=res.get(k, ""))
        wb.save(xlsx_path)
    except ImportError:
        pass
    
    # JSON
    json_path = DATA_DIR / f"{pais.lower()}_extraido.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)


def generar_visor_datos():
    """Combina todos los países en datos.js"""
    todos = []
    
    for idx_file in sorted(DATA_DIR.glob("*_extraido.json")):
        pais = idx_file.stem.split("_")[0].upper()
        try:
            with open(idx_file) as f:
                data = json.load(f)
            for d in data:
                d.setdefault("provincia", pais)
                d.setdefault("operador", "")
                d.setdefault("linea", "")
                d.setdefault("pk", "")
                d.setdefault("hora", "")
                d.setdefault("latitud", 48.0)
                d.setdefault("longitud", 10.0)
                d.setdefault("resumen", "")
                d.setdefault("fecha", "2000-01-01")
                todos.append(d)
        except:
            pass
    
    # Ordenar por fecha
    todos.sort(key=lambda x: x.get("fecha", ""))
    
    # Generar datos.js
    js_content = "window.DATOS = " + json.dumps(todos, ensure_ascii=False, indent=2) + ";"
    visor_path = VISOR_DIR / "datos.js"
    visor_path.write_text(js_content, encoding="utf-8")
    
    return len(todos)


# === MAIN ===
def main():
    print("=" * 60)
    print(f"🚆 ERAVisor Pipeline Europa v2 — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)
    
    # Cargar estado
    state = cargar_estado()
    
    # Cargar índices
    indices = cargar_indices()
    if not indices:
        print("❌ No hay índices disponibles")
        return
    
    total_reports = sum(len(r) for r in indices.values())
    print(f"📡 {len(indices)} países indexados ({total_reports} informes total)")
    
    # Fase 1: Descargar PDFs (batch de 10)
    print(f"\n📥 Fase 1: Descarga de PDFs...")
    desc, pend = descargar_pendientes(indices, state, batch_size=10)
    if desc > 0:
        print(f"  ✅ {desc} PDFs descargados ({pend} pendientes)")
    else:
        print(f"  ✅ Sin pendientes")
    
    guardar_estado(state)
    print(f"  💾 Estado guardado: {sum(len(v) for v in state.get('downloaded', {}).values())} descargados, {sum(len(v) for v in state.get('extracted', {}).values())} extraídos")
    
    # Fase 2: Extraer con LLM (batch de 5)
    print(f"\n🤖 Fase 2: Extracción con LLM...")
    ext, pend_ext = extraer_pendientes(indices, state, batch_size=5)
    if ext > 0:
        print(f"  ✅ {ext} accidentes extraídos ({pend_ext - ext} pendientes)")
    else:
        print(f"  ✅ Sin pendientes")
    
    guardar_estado(state)
    
    # Fase 3: Actualizar visor
    print(f"\n🌐 Fase 3: Actualizando visor...")
    total_vis = generar_visor_datos()
    print(f"  🌐 Visor: {total_vis} registros")
    
    print(f"\n{'='*60}")
    print(f"✅ RESUMEN:")
    print(f"   📥 Descargas: {desc}")
    print(f"   🤖 Extracciones: {ext}")
    print(f"   🌐 Visor total: {total_vis} registros")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
