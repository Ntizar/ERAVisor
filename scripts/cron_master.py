#!/usr/bin/env python3
"""
ERAVisor — Cron Master: indexa todos los países ERA, descarga PDFs nuevos,
extrae datos con qwen3.6, actualiza dataset y visor.

Se ejecuta semanalmente o bajo demanda.
"""
import json, csv, os, re, sys, time
from pathlib import Path
from datetime import datetime
import requests, PyPDF2, openpyxl
from openpyxl.styles import Font, PatternFill

BASE = Path("/root/workspace/ERAVisor")
PDF_DIR = BASE / "pdfs"
DATA_DIR = BASE / "data"
VISOR_DIR = BASE / "visor"
SCRIPT_DIR = BASE / "scripts"

NAN_KEY = os.environ.get("NAN_API", "")
LLM_URL = "https://api.nan.builders/v1/chat/completions"
LLM_MODEL = "qwen3.6"  # <-- qwen3.6 para clasificación

# === TAXONOMÍA COMPLETA Anexo III ===
TAX_SUCESOS = {
    "1":"Accidente","1.1":"Colisión tren-vehículo","1.1.1":"Frontal","1.1.2":"Alcance","1.1.3":"Lateral",
    "1.2":"Colisión obstáculo","1.2.1":"Elementos tren","1.2.2":"Superestructura","1.2.3":"Infraestructura",
    "1.2.4":"Protección itinerarios","1.2.5":"Animales","1.2.6":"Árboles","1.2.7":"Maq. mantenimiento","1.2.8":"Otros",
    "1.3":"Descarrilamiento","1.3.1":"Plena vía","1.3.2":"Estaciones",
    "1.4":"Accidente paso nivel","1.4.1":"PN Usuarios","1.4.1.1":"PN Vehículo","1.4.1.2":"PN Peatón","1.4.2":"PN Objetos",
    "1.5":"Atropello persona","1.5.1":"Arrollamiento","1.5.1.1":"Plena vía","1.5.1.2":"Estaciones",
    "1.5.1.2.1":"Cruce andenes","1.5.1.2.2":"Estaciones otros","1.5.2":"Caídas","1.5.2.1":"En marcha","1.5.2.2":"Subir/bajar",
    "1.5.3":"Otros personas","1.5.3.1":"Movimiento vehículo","1.5.3.2":"Proyección objetos","1.5.3.3":"Otros",
    "1.6":"Incendio/explosión",
    "1.7":"Otros accidentes","1.7.1":"Proc. emergencia","1.7.2":"Maniobras","1.7.3":"Cambio ancho",
    "1.7.4":"Vía bloqueada","1.7.5":"Mercancías peligrosas","1.7.6":"Fallo cargamento","1.7.7":"Otros",
    "2":"Incidente","2.1":"Precursor","2.1.1":"Rotura carril","2.1.2":"Deformación vía",
    "2.1.3":"Fallo señalización","2.1.3.1":"Señalización infra","2.1.3.2":"Señalización MR",
    "2.1.4":"SPAD sobrepasando","2.1.4.1":"Conato colisión","2.1.4.2":"SPAD otros",
    "2.1.5":"SPAD sin sobrepasar","2.1.6":"Rueda rota","2.1.7":"Eje roto",
    "2.2":"Otros precursores","2.2.1":"Incidentes operacionales",
    "2.2.1.1":"Exceso velocidad","2.2.1.2":"Marcha no autorizada","2.2.1.3":"Incumplimiento bloqueo","2.2.1.4":"Otros incumplimientos",
    "2.2.2":"Escape material","2.2.2.1":"Conato colisión","2.2.2.2":"Escape otros",
    "2.3":"Otros incidentes","2.3.1":"Componentes vehículo","2.3.2":"Fallo cargamento","2.3.3":"Incidente PN",
    "2.3.4":"Talonamiento","2.3.5":"Conato incendio","2.3.6":"Mercancías peligrosas",
    "2.3.7":"Incidente obstáculo gálibo","2.3.7.1":"Elementos tren","2.3.7.2":"Superestructura","2.3.7.3":"Infraestructura",
    "2.3.7.4":"Protección itinerarios","2.3.7.5":"Animales","2.3.7.6":"Árboles","2.3.7.7":"Otros","2.3.8":"Otras incidencias",
    "3":"Suicidio","3.1":"Suicidio","3.2":"Intento suicidio",
}
TAX_CAUSAS = {
    "1":"Ferrocarril","1.1":"Factor humano","1.1.1":"Señales","1.1.2":"Bloqueo","1.1.3":"Itinerario",
    "1.1.4":"Formación tren","1.1.5":"Conducción","1.1.6":"Mantenimiento","1.1.7":"Trabajos vía",
    "1.1.8":"Maniobras","1.1.9":"Prescripciones","1.1.10":"Otros FH",
    "1.2":"Fallo técnico","1.2.1":"Fallo MR","1.2.1.1":"Rodadura","1.2.1.2":"Suspensión","1.2.1.3":"Freno",
    "1.2.1.4":"Tracción-choque","1.2.1.5":"Gálibo","1.2.1.6":"Pantógrafo","1.2.1.7":"Puertas",
    "1.2.1.8":"Seguridad embarcada","1.2.1.9":"Motor eléctrico","1.2.1.10":"Motor combustión",
    "1.2.1.11":"Elemento caído","1.2.1.12":"Otros MR",
    "1.2.2":"Fallo instalaciones","1.2.2.1":"Infraestructura","1.2.2.2":"Vía","1.2.2.3":"Carril",
    "1.2.2.4":"Aparato vía","1.2.2.5":"Seguridad","1.2.2.6":"Electrificación","1.2.2.7":"Otras instalaciones",
    "2":"Usuarios/entorno","2.1":"Usuarios FC","2.1.1":"Viajeros","2.1.2":"Cliente mercancías","2.1.3":"Otros usuarios",
    "2.2":"Condiciones entorno","2.2.1":"Meteorológicas","2.2.2":"Medio ambiente",
    "2.3":"Otros","2.3.1":"Extraños","2.3.2":"Intrusos","2.3.3":"Usuario PN","2.3.4":"Personas andén","2.3.5":"Personas fuera andén",
    "2.4":"Sin identificar",
}

# === PASO 1: Indexar país ===
def indexar_pais(pais, nombre_pais):
    """Indexa informes ERA de un país"""
    idx_file = DATA_DIR / f"{pais.lower()}-investigations-index.json"
    
    print(f"  📡 ERA: {pais} ({nombre_pais})...", end=" ", flush=True)
    
    try:
        r = requests.get(
            f"https://www.era.europa.eu/era-folder/accident-investigation-reports",
            params={"f[0]": f"country:{pais}"},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=30
        )
        # Si no responde bien, devolver índice existente
        if r.status_code != 200:
            existing = json.load(open(idx_file)) if idx_file.exists() else {"reports":[]}
            print(f"⚠️ HTTP {r.status_code}, usando cache ({len(existing['reports'])} reports)")
            return existing["reports"]
    except:
        existing = json.load(open(idx_file)) if idx_file.exists() else {"reports":[]}
        print(f"⚠️ error red, usando cache ({len(existing['reports'])} reports)")
        return existing["reports"]
    
    # Parsear página — buscamos enlaces a PDFs con el formato ERA
    reports = []
    # Buscar patrones de PDF en el HTML
    pdfs = re.findall(r'(/system/files/[^"\']+\.pdf)', r.text)
    seen = set()
    for p in pdfs:
        url = f"https://www.era.europa.eu{p}" if p.startswith('/') else p
        if url in seen:
            continue
        seen.add(url)
        # Extraer año y título del nombre del archivo
        fname = url.split('/')[-1]
        year = datetime.now().year
        m = re.search(r'(\d{4})', fname)
        if m:
            year = int(m.group(1))
        reports.append({
            "title": fname,
            "pdf_url": url,
            "year": year,
            "country": pais
        })
    
    if reports:
        with open(idx_file, "w") as f:
            json.dump({"source": "ERA", "total_reports": len(reports), "reports": reports, "fetched_at": datetime.now().isoformat()}, f, indent=2)
        print(f"✅ {len(reports)} informes")
    else:
        existing = json.load(open(idx_file)) if idx_file.exists() else {"reports":[]}
        reports = existing["reports"]
        print(f"⚠️ sin parsear, usando cache ({len(reports)} reports)")
    
    return reports

# === PASO 2: Descargar PDFs nuevos ===
def descargar_nuevos(pais, reports):
    """Descarga PDFs que no estén ya en disco"""
    pais_dir = PDF_DIR / pais
    pais_dir.mkdir(parents=True, exist_ok=True)
    
    nuevos = 0
    total = len(reports)
    
    for i, r in enumerate(reports):
        title = r["title"]
        pdf_url = r["pdf_url"]
        fname = f"{pais}_{r['year']}_{title}"
        path = pais_dir / fname
        
        if path.exists() and path.stat().st_size > 2000:
            continue
        
        for intento in range(3):
            try:
                resp = requests.get(pdf_url, headers={
                    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0"
                }, timeout=60, allow_redirects=True)
                if resp.status_code == 200 and resp.content[:5] == b"%PDF-":
                    path.write_bytes(resp.content)
                    nuevos += 1
                    print(f"    📥 [{i+1}/{total}] {title[:55]}")
                    break
                elif resp.status_code == 429:
                    time.sleep(3 * (intento+1))
                else:
                    break
            except:
                time.sleep(2 * (intento+1))
    
    return nuevos

# === PASO 3: Extraer con regex + qwen3.6 ===
def extraer_pdf(path_pdf, title, pais, year, pdf_url):
    """Extrae datos de un PDF con regex + qwen3.6 para clasificación"""
    try:
        reader = PyPDF2.PdfReader(open(path_pdf, "rb"))
        paginas = [p.extract_text() or "" for p in reader.pages]
        texto = "\n\n".join(paginas)
        npag = len(reader.pages)
    except Exception as e:
        print(f"    ❌ Error PDF: {e}")
        return None
    
    txt_head = texto[:5000]
    txt_upper = txt_head.upper()
    
    datos = {
        "pais": pais,
        "titulo_informe": title,
        "informe_publico_url": pdf_url,
        "fecha_procesamiento": datetime.now().isoformat(),
        "fuente_datos": "ERA",
        "organismo_investigador": "CIAF" if pais == "ES" else pais,
        "estado_investigacion": "Completado",
        "num_paginas": npag,
        "fallecidos": 0, "heridos_graves": 0, "heridos_leves": 0,
        "latitud": 40.0, "longitud": -3.0,
    }
    
    # Extraer fecha
    m = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', txt_head)
    if m and 1900 < int(m.group(3)) < 2030:
        d, mo, y = m.group(1).zfill(2), m.group(2).zfill(2), m.group(3)
        datos["fecha"] = f"{y}-{mo}-{d}"
    else:
        datos["fecha"] = f"{year}-01-01"
    
    # Generar ID
    stem = Path(title).stem[:20]
    datos["id_accidente"] = f"{pais}-{year}-{stem}"
    
    # Clasificación con qwen3.6
    print(f"    🤖 qwen3.6 clasificando...", end=" ", flush=True)
    llm_data = clasificar_con_qwen(texto[:8000], pais)
    if llm_data:
        for k in ["suceso_codigo", "suceso_desc", "causa_codigo", "causa_desc", "resumen",
                   "provincia", "municipio", "operador", "linea", "pk", "hora",
                   "fallecidos", "heridos_graves", "heridos_leves"]:
            if k in llm_data and llm_data[k] not in (None, "", "null"):
                datos[k] = llm_data[k]
        print("✅")
    else:
        print("⚠️ fallback regex")
        # Fallback: regex simple para tipo
        if "DESCARRILAMIENTO" in txt_upper:
            datos["suceso_codigo"] = "1.3"
            datos["suceso_desc"] = TAX_SUCESOS["1.3"]
        elif "COLISI" in txt_upper:
            datos["suceso_codigo"] = "1.1"
            datos["suceso_desc"] = TAX_SUCESOS["1.1"]
        elif "PASO A NIVEL" in txt_upper or "PASO A NIVEL" in txt_upper:
            datos["suceso_codigo"] = "1.4"
            datos["suceso_desc"] = TAX_SUCESOS["1.4"]
        elif "ARROLLAMIENTO" in txt_upper or "ATROPELLO" in txt_upper:
            datos["suceso_codigo"] = "1.5"
            datos["suceso_desc"] = TAX_SUCESOS["1.5"]
        elif "INCENDIO" in txt_upper:
            datos["suceso_codigo"] = "1.6"
            datos["suceso_desc"] = TAX_SUCESOS["1.6"]
        elif "SUICIDIO" in txt_upper:
            datos["suceso_codigo"] = "3"
            datos["suceso_desc"] = TAX_SUCESOS["3"]
        else:
            datos["suceso_codigo"] = "1.7.7"
            datos["suceso_desc"] = TAX_SUCESOS["1.7.7"]
        datos["causa_codigo"] = "2.4"
        datos["causa_desc"] = TAX_CAUSAS["2.4"]
    
    # Resumen
    lines = [l.strip() for l in texto.split('\n') if l.strip() and len(l) > 50]
    if lines:
        datos["resumen"] = datos.get("resumen", lines[0][:400])
    
    return datos


def clasificar_con_qwen(texto, pais):
    """Clasifica el accidente usando qwen3.6"""
    prompt = f"""Eres un experto en seguridad ferroviaria. Analiza este informe de accidente y clasifícalo según el Anexo III del RD 929/2020.

INFORME (país: {pais}):
{texto[:6000]}

CLASIFICACIÓN DISPONIBLE (código - descripción):
SUCESOS: {', '.join(f'{k}={v}' for k,v in sorted(TAX_SUCESOS.items()))}

CAUSAS: {', '.join(f'{k}={v}' for k,v in sorted(TAX_CAUSAS.items()))}

Responde SOLO con este JSON exacto (sin texto adicional, sin markdown):
{{"suceso_codigo":"1.1.1","suceso_desc":"Colisión frontal","causa_codigo":"1.1.5","causa_desc":"FH - Conducción","provincia":"Madrid","municipio":"","operador":"Renfe","linea":"Madrid-Hendaya","fecha":"2024-01-15","hora":"10:30","pk":"73+500","fallecidos":0,"heridos_graves":2,"heridos_leves":5,"resumen":"Descripción breve del accidente"}}

Usa código de suceso y causa de la lista. Si no puedes determinar, usa "1.7.7" para suceso y "2.4" para causa. Responde SÓLO el JSON."""
    
    try:
        r = requests.post(LLM_URL, headers={
            "Authorization": f"Bearer {NAN_KEY}", "Content-Type": "application/json"
        }, json={
            "model": LLM_MODEL,
            "messages": [
                {"role": "system", "content": "Eres un experto en seguridad ferroviaria. Responde SOLO con JSON válido."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.05,
            "max_tokens": 1200
        }, timeout=180)
        
        if r.status_code == 200:
            text = r.json()["choices"][0]["message"]["content"]
            m = re.search(r'\{.*\}', text, re.DOTALL)
            if m:
                return json.loads(m.group())
    except Exception as e:
        print(f"⚠️ LLM error: {e}", end=" ")
    
    return None

# === PASO 4: Exportar a CSV/Excel ===
def exportar(resultados, pais):
    if not resultados:
        return
    
    keys = sorted(set().union(*(r.keys() for r in resultados)))
    
    # CSV
    csv_path = DATA_DIR / f"{pais.lower()}_accidentes.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        w.writerows(resultados)
    
    # Excel
    xlsx_path = DATA_DIR / f"{pais.lower()}_accidentes.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Accidentes {pais}"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for c, k in enumerate(keys, 1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.font = hf
        cell.fill = hfill
    for ri, res in enumerate(resultados, 2):
        for c, k in enumerate(keys, 1):
            ws.cell(row=ri, column=c, value=res.get(k, ""))
    wb.save(xlsx_path)
    
    # JSON
    json_path = DATA_DIR / f"{pais.lower()}_extraido.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)
    
    print(f"  📄 CSV/Excel/JSON: {len(resultados)} registros")

# === PASO 5: Generar datos.js para visor ===
def generar_visor_datos():
    """Combina todos los países en datos.js"""
    todos = []
    
    for idx_file in sorted(DATA_DIR.glob("*_extraido.json")):
        pais = idx_file.stem.split("_")[0].upper()
        try:
            with open(idx_file) as f:
                data = json.load(f)
            for d in data:
                # Asegurar campos mínimos para el visor
                d.setdefault("provincia", pais)
                d.setdefault("operador", "")
                d.setdefault("linea", "")
                d.setdefault("pk", "")
                d.setdefault("tipo_trafico", "")
                d.setdefault("tipo_linea", "")
                d.setdefault("paso_nivel", "")
                d.setdefault("estacion", "")
                d.setdefault("hora", "")
                d.setdefault("trenes", 1)
                d.setdefault("causas_directas", "")
                d.setdefault("factores_contribuyentes", "")
                d.setdefault("recomendaciones", "")
                d.setdefault("latitud", 40.0)
                d.setdefault("longitud", -3.0)
                d.setdefault("resumen", "")
                d.setdefault("fecha", "2000-01-01")
                todos.append(d)
        except:
            print(f"  ⚠️ Error cargando {idx_file.name}")
    
    # Generar datos.js
    js_content = "window.DATOS = " + json.dumps(todos, ensure_ascii=False, indent=2) + ";"
    visor_path = VISOR_DIR / "datos.js"
    visor_path.write_text(js_content, encoding="utf-8")
    print(f"  🌐 Visor datos.js: {len(todos)} registros totales")

# === MAIN ===
def main():
    print("=" * 60)
    print(f"🚆 ERAVisor Cron — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"🤖 Modelo: {LLM_MODEL}")
    print("=" * 60)
    
    PAISES = [
        ("ES", "España"), ("DE", "Alemania"), ("FR", "Francia"),
        ("IT", "Italia"), ("UK", "Reino Unido"), ("NL", "Países Bajos"),
        ("AT", "Austria"), ("BE", "Bélgica"), ("BG", "Bulgaria"),
        ("CH", "Suiza"), ("CZ", "República Checa"), ("DK", "Dinamarca"),
        ("EE", "Estonia"), ("FI", "Finlandia"), ("GR", "Grecia"),
        ("HR", "Croacia"), ("HU", "Hungría"), ("IE", "Irlanda"),
        ("LT", "Lituania"), ("LU", "Luxemburgo"), ("LV", "Letonia"),
        ("NO", "Noruega"), ("PL", "Polonia"), ("PT", "Portugal"),
        ("RO", "Rumanía"), ("SE", "Suecia"), ("SI", "Eslovenia"),
        ("SK", "Eslovaquia"), ("RS", "Serbia"),
    ]
    
    total_indexados = 0
    total_nuevos = 0
    total_extraidos = 0
    
    for pais, nombre in PAISES:
        print(f"\n📍 {pais} — {nombre}")
        
        # 1. Indexar
        reports = indexar_pais(pais, nombre)
        total_indexados += len(reports)
        
        if not reports:
            continue
        
        # 2. Descargar nuevos
        nuevos = descargar_nuevos(pais, reports)
        total_nuevos += nuevos
        if nuevos:
            print(f"  📥 {nuevos} nuevos PDFs descargados")
        
        # 3. Extraer solo los que no tienen extracción
        pais_dir = PDF_DIR / pais
        extraido_file = DATA_DIR / f"{pais.lower()}_extraido.json"
        
        # Cargar extraídos previos
        existentes = {}
        if extraido_file.exists():
            for d in json.load(open(extraido_file)):
                if "titulo_informe" in d:
                    existentes[d["titulo_informe"]] = d
        
        # Ver cuáles PDFs están descargados pero no extraídos
        pendientes = []
        for r in reports:
            fname = f"{pais}_{r['year']}_{r['title']}"
            path = pais_dir / fname
            if path.exists() and r["title"] not in existentes:
                pendientes.append((r, path))
        
        if not pendientes:
            print(f"  ✅ Todos extraídos ({len(existentes)} previos)")
            continue
        
        print(f"  🔄 {len(pendientes)} pendientes de extraer...")
        resultados = list(existentes.values())
        
        for i, (r, path) in enumerate(pendientes):
            print(f"  [{i+1}/{len(pendientes)}] {r['title'][:55]}", end=" ", flush=True)
            d = extraer_pdf(path, r["title"], pais, r["year"], r["pdf_url"])
            if d:
                resultados.append(d)
                total_extraidos += 1
            # Delay entre llamadas LLM para no saturar
            time.sleep(0.5)
        
        if resultados:
            exportar(resultados, pais)
    
    # 5. Generar visor combinado
    print(f"\n{'='*60}")
    print(f"🌐 Generando visor combinado...")
    generar_visor_datos()
    
    print(f"\n{'='*60}")
    print(f"✅ RESUMEN:")
    print(f"   📡 Países indexados: {len(PAISES)}")
    print(f"   📦 Informes totales: {total_indexados}")
    print(f"   📥 Nuevas descargas: {total_nuevos}")
    print(f"   🤖 Nuevas extracciones: {total_extraidos}")
    print(f"   🌐 Visor datos.js actualizado")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()