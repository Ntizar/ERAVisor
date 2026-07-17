#!/usr/bin/env python3
"""
ERAVisor Pipeline — Descarga + Extracción Estructurada de PDFs ERA
Uso: python3 extract_pipeline.py --pais ES [--max 10] [--samples 5]
"""

import json, csv, os, sys, time, re, argparse
from datetime import datetime
from pathlib import Path
import requests, PyPDF2, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = Path("/root/workspace/ERAVisor")
PDF_DIR = BASE_DIR / "pdfs"
DATA_DIR = BASE_DIR / "data"
NAN_API_KEY = os.environ.get("NAN_API", "")
NAN_API_URL = "https://api.nan.builders/v1/chat/completions"
LLM_MODEL = "deepseek-v4-flash"
DOWNLOAD_DELAY = 2.0
MAX_RETRIES = 5


def cargar_taxonomia_sucesos():
    return {
        "1": "Accidente", "1.1": "Colisión tren con vehículo ferroviario",
        "1.1.1": "Frontal", "1.1.2": "Alcance", "1.1.3": "Lateral",
        "1.2": "Colisión tren con obstáculo en gálibo",
        "1.2.1": "Elementos de tren", "1.2.2": "Superestructura",
        "1.2.3": "Infraestructura", "1.2.4": "Protección itinerarios",
        "1.2.5": "Animales", "1.2.6": "Árboles", "1.2.7": "Maquinaria mantenimiento",
        "1.2.8": "Otros", "1.3": "Descarrilamiento de tren",
        "1.3.1": "Plena vía", "1.3.2": "Estaciones",
        "1.4": "Accidente en paso a nivel",
        "1.4.1": "Usuarios", "1.4.1.1": "Vehículos carretera", "1.4.1.2": "Peatones",
        "1.4.2": "Objetos",
        "1.5": "Accidente a persona por material rodante",
        "1.5.1": "Arrollamiento", "1.5.1.1": "Plena vía", "1.5.1.2": "Estaciones",
        "1.5.1.2.1": "Cruce andenes", "1.5.1.2.2": "Otros",
        "1.5.2": "Caídas", "1.5.2.1": "En marcha", "1.5.2.2": "Subir/bajar",
        "1.5.3": "Otros accidentes personas",
        "1.5.3.1": "Movimiento vehículo", "1.5.3.2": "Proyección objetos",
        "1.5.3.3": "Otros",
        "1.6": "Incendio/explosión en material rodante",
        "1.7": "Otros accidentes",
        "1.7.1": "Colisión por procedimiento emergencia",
        "1.7.2": "Colisión/descarrilamiento maniobras",
        "1.7.3": "Descarrilamiento cambio ancho",
        "1.7.4": "Colisión vía bloqueada/mantenimiento",
        "1.7.5": "Accidente mercancías peligrosas",
        "1.7.6": "Fallo cargamento", "1.7.7": "Otros",
        "2": "Incidente", "2.1": "Precursor",
        "2.1.1": "Rotura carril", "2.1.2": "Deformación vía",
        "2.1.3": "Fallo señalización", "2.1.3.1": "Infraestructura",
        "2.1.3.2": "Material rodante",
        "2.1.4": "Señal rebasada peligro (con rebase)",
        "2.1.4.1": "Conato colisión", "2.1.4.2": "Otros",
        "2.1.5": "Señal rebasada peligro (sin rebase)",
        "2.1.6": "Rueda rota", "2.1.7": "Eje roto",
        "2.2": "Otros precursores",
        "2.2.1": "Incidentes operacionales",
        "2.2.1.1": "Exceso velocidad", "2.2.1.2": "Marcha no autorizada",
        "2.2.1.3": "Incumplimiento bloqueo",
        "2.2.1.4": "Otros incumplimientos seguridad",
        "2.2.2": "Escape material", "2.2.2.1": "Conato colisión", "2.2.2.2": "Otros",
        "2.3": "Otros incidentes",
        "2.3.1": "Componentes vehículo", "2.3.2": "Fallo cargamento",
        "2.3.3": "Incidente paso nivel", "2.3.4": "Talonamiento",
        "2.3.5": "Conato incendio/explosión",
        "2.3.6": "Incidente mercancías peligrosas",
        "2.3.7": "Incidente colisión obstáculo",
        "2.3.7.1": "Elementos tren", "2.3.7.2": "Superestructura",
        "2.3.7.3": "Infraestructura", "2.3.7.4": "Protección itinerarios",
        "2.3.7.5": "Animales", "2.3.7.6": "Árboles", "2.3.7.7": "Otros",
        "2.3.8": "Otras incidencias",
        "3": "Suicidio", "3.1": "Suicidio", "3.2": "Intento suicidio",
    }


def cargar_taxonomia_causas():
    return {
        "1": "Ferrocarril", "1.1": "Factor humano",
        "1.1.1": "Señales", "1.1.2": "Bloqueo", "1.1.3": "Itinerario",
        "1.1.4": "Formación tren", "1.1.5": "Conducción",
        "1.1.6": "Mantenimiento MR e instalaciones",
        "1.1.7": "Trabajos vía", "1.1.8": "Maniobras",
        "1.1.9": "Prescripciones circulación", "1.1.10": "Otros",
        "1.2": "Fallo técnico",
        "1.2.1": "Fallo material rodante",
        "1.2.1.1": "Rodadura", "1.2.1.2": "Suspensión", "1.2.1.3": "Freno",
        "1.2.1.4": "Tracción-choque", "1.2.1.5": "Gálibo", "1.2.1.6": "Pantógrafo",
        "1.2.1.7": "Puertas", "1.2.1.8": "Seguridad embarcadas",
        "1.2.1.9": "Motor eléctrico", "1.2.1.10": "Motor combustión",
        "1.2.1.11": "Elemento caído vehículo", "1.2.1.12": "Otros",
        "1.2.2": "Fallo instalaciones",
        "1.2.2.1": "Infraestructura", "1.2.2.2": "Vía", "1.2.2.3": "Carril",
        "1.2.2.4": "Aparato vía", "1.2.2.5": "Seguridad", "1.2.2.6": "Electrificación",
        "1.2.2.7": "Otros",
        "2": "Usuarios/entorno/otros", "2.1": "Usuarios ferrocarril",
        "2.1.1": "Viajeros", "2.1.2": "Cliente mercancías", "2.1.3": "Otros",
        "2.2": "Condiciones entorno",
        "2.2.1": "Meteorológicas", "2.2.2": "Medio ambiente",
        "2.3": "Otros", "2.3.1": "Extraños", "2.3.2": "Intrusos",
        "2.3.3": "Usuarios paso nivel", "2.3.4": "Personas andén",
        "2.3.5": "Personas fuera andén", "2.4": "Sin identificar",
    }


def extraer_texto_pdf(ruta):
    try:
        with open(ruta, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            paginas = []
            for i, page in enumerate(reader.pages):
                paginas.append(f"--- PAGINA {i+1} ---\n{page.extract_text()}")
            return "\n\n".join(paginas), len(reader.pages)
    except Exception as e:
        return f"[ERROR: {e}]", 0


def descargar_pdf(url, destino):
    if destino.exists() and destino.stat().st_size > 1000:
        print(f"  OK ya existe: {destino.name}")
        return True
    for intento in range(MAX_RETRIES):
        try:
            h = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0"}
            r = requests.get(url, headers=h, timeout=60, allow_redirects=True)
            if r.status_code == 200 and r.content[:5] == b"%PDF-":
                destino.write_bytes(r.content)
                print(f"  OK descargado: {destino.name} ({len(r.content)//1024}KB)")
                return True
            elif r.status_code == 429:
                e = DOWNLOAD_DELAY * (intento + 1) * 2
                print(f"  429, esperando {e}s...")
                time.sleep(e)
            elif r.status_code == 404:
                print(f"  404: {url[:80]}")
                return False
            else:
                print(f"  HTTP {r.status_code}: {url[:80]}")
                time.sleep(DOWNLOAD_DELAY)
        except Exception as e:
            print(f"  Error descarga: {e}")
            time.sleep(DOWNLOAD_DELAY * (intento + 1))
    return False


def llamar_llm(prompt, sistema="Eres un experto en seguridad ferroviaria."):
    if not NAN_API_KEY:
        return None
    try:
        r = requests.post(NAN_API_URL, headers={
            "Authorization": f"Bearer {NAN_API_KEY}", "Content-Type": "application/json"
        }, json={
            "model": LLM_MODEL, "messages": [
                {"role": "system", "content": sistema},
                {"role": "user", "content": prompt}
            ], "temperature": 0.1, "max_tokens": 2000
        }, timeout=120)
        if r.status_code == 200:
            return r.json()["choices"][0]["message"]["content"]
        print(f"  LLM error {r.status_code}: {r.text[:200]}")
        return None
    except Exception as e:
        print(f"  LLM error: {e}")
        return None


def extraer_basicos(texto, nombre_archivo, anio, pais):
    res = {"id_accidente": f"ERA-{pais}-{anio}-{Path(nombre_archivo).stem[:20]}",
           "pais": pais, "informe_publico_url": ""}
    # Fecha
    m = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', texto[:5000])
    if m:
        res["fecha"] = f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    if "fecha" not in res:
        res["fecha"] = f"{anio}-01-01"
    # Hora
    m = re.search(r'(?:a las|sobre las)\s*(\d{1,2})[:h](\d{2})?', texto[:3000], re.I)
    if m:
        res["hora"] = f"{m.group(1).zfill(2)}:{m.group(2) or '00'}"
    # Provincia
    m = re.search(r'(?:provincia de|término municipal de)\s+([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+)*)', texto[:5000], re.I)
    if m:
        res["provincia"] = m.group(1).strip()
    return res


def construir_prompt(texto, basicos, tax_s, tax_c):
    txt = texto[:12000]
    ts = "\n".join(f"  {k}: {v}" for k, v in sorted(tax_s.items()))
    tc = "\n".join(f"  {k}: {v}" for k, v in sorted(tax_c.items()))
    return f"""A partir del informe de accidente ferroviario extrae datos estructurados.

INFORME:
{txt}

TAXONOMIA SUCESOS:
{ts}

TAXONOMIA CAUSAS:
{tc}

Responde SOLO con JSON:
{{"fecha":"YYYY-MM-DD","hora":"HH:MM","provincia":"","municipio":"","operador":"","infraestructura":"","linea":"","pk":"","tipo_trafico":"","tipo_linea":"","paso_nivel":false,"estacion":false,"suceso_codigo":"","suceso_desc":"","causa_codigo":"","causa_desc":"","fallecidos":0,"heridos_graves":0,"heridos_leves":0,"trenes":1,"resumen":"","causas_directas":"","factores_contribuyentes":"","recomendaciones":""}}

Usa los codigos de las taxonomias. null si no disponible."""


def procesar_pdf(report, pais_dir):
    title, url, year, country = report["title"], report["pdf_url"], report["year"], report["country"]
    pdf_name = f"{country}_{year}_{Path(title).name}"
    pdf_path = pais_dir / pdf_name
    print(f"\nProcesando: {title}")

    if not descargar_pdf(url, pdf_path):
        return None
    texto, npag = extraer_texto_pdf(pdf_path)
    print(f"  {npag} paginas")
    if npag == 0:
        return None

    basicos = extraer_basicos(texto, title, year, country)
    basicos["informe_publico_url"] = url

    prompt = construir_prompt(texto, basicos, cargar_taxonomia_sucesos(), cargar_taxonomia_causas())
    resp = llamar_llm(prompt)
    if resp:
        try:
            jm = re.search(r'\{.*\}', resp, re.DOTALL)
            if jm:
                datos = json.loads(jm.group())
                for k, v in datos.items():
                    if v is not None:
                        basicos[k] = v
                print(f"  OK extraccion LLM completa")
        except json.JSONDecodeError:
            print(f"  WARN JSON invalido, usando regex")
    basicos["titulo_informe"] = title
    basicos["num_paginas"] = npag
    basicos["fecha_procesamiento"] = datetime.now().isoformat()
    return basicos


def generar_excel(resultados, ruta):
    if not resultados:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accidentes ERA"
    keys = sorted(set().union(*(r.keys() for r in resultados)))
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for c, k in enumerate(keys, 1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.font = hf
        cell.fill = hfill
    for r_idx, res in enumerate(resultados, 2):
        for c, k in enumerate(keys, 1):
            ws.cell(row=r_idx, column=c, value=res.get(k, ""))
    wb.save(ruta)
    print(f"Excel: {ruta} ({len(resultados)} filas, {len(keys)} cols)")


def generar_csv(resultados, ruta):
    if not resultados:
        return
    keys = sorted(set().union(*(r.keys() for r in resultados)))
    with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        w.writerows(resultados)
    print(f"CSV: {ruta} ({len(resultados)} filas)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pais", default="ES")
    parser.add_argument("--max", type=int, default=0)
    parser.add_argument("--samples", type=int, default=5)
    parser.add_argument("--skip-llm", action="store_true")
    args = parser.parse_args()

    index_path = DATA_DIR / f"{args.pais.lower()}-investigations-index.json"
    if not index_path.exists():
        print(f"Index no encontrado: {index_path}")
        return
    with open(index_path) as f:
        index = json.load(f)
    reports = index["reports"]
    print(f"Total indices: {len(reports)}")

    pais_dir = PDF_DIR / args.pais
    pais_dir.mkdir(parents=True, exist_ok=True)

    if args.samples:
        # Muestra representativa por decadas
        anios = sorted(set(r["year"] for r in reports))
        paso = max(1, len(anios) // args.samples)
        anios_sel = anios[::paso][:args.samples]
        reports_proc = [r for r in reports if r["year"] in anios_sel]
    elif args.max:
        reports_proc = reports[:args.max]
    else:
        reports_proc = reports

    print(f"A procesar: {len(reports_proc)} informes")
    resultados = []

    for i, report in enumerate(reports_proc):
        print(f"\n[{i+1}/{len(reports_proc)}]")
        r = procesar_pdf(report, pais_dir)
        if r:
            resultados.append(r)
        time.sleep(1.0)

    if resultados:
        json_path = DATA_DIR / f"{args.pais.lower()}_extraido.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(resultados, f, ensure_ascii=False, indent=2)

        csv_path = DATA_DIR / f"{args.pais.lower()}_accidentes.csv"
        generar_csv(resultados, csv_path)

        xlsx_path = DATA_DIR / f"{args.pais.lower()}_accidentes.xlsx"
        generar_excel(resultados, xlsx_path)

    print(f"\nFinalizado. {len(resultados)}/{len(reports_proc)} procesados OK")


if __name__ == "__main__":
    main()